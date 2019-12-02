from django.shortcuts import render

# Create your views here.

from rest_framework.pagination import PageNumberPagination
from ipdata import models
from .models import BusinessUnit,IpInfo,Environment,RecordingTime,AllData
import datetime
from rest_framework import serializers
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.colors import GREEN, BLACK
from django.http import FileResponse
import os



class UsersSerializer(serializers.ModelSerializer):
    environment_set = serializers.StringRelatedField(many=True)
    class Meta:
        model = models.BusinessUnit
        fields = "__all__"				# 查询表内所有的内容
        depth = 2					# 设置连表的深度，如果被连表还有连表 可以设置为2


class IpinfoSerializer(serializers.ModelSerializer):
    times = serializers.StringRelatedField(label='时间')
    servertype = serializers.CharField(source="get_servertype_display")
    type = serializers.CharField(source="get_type_display")
    class Meta:
        model = IpInfo
        fields = "__all__"				# 查询表内所有的内容


class AlldataSerializer(serializers.ModelSerializer):

    time = serializers.StringRelatedField(label='时间')
    class Meta:
        model = AllData
        fields = "__all__"				# 查询表内所有的内容



class Base_info(APIView):
    def get(self,request,*args,**kwargs):
        ret = models.BusinessUnit.objects.all()
        ser = UsersSerializer(instance=ret,many=True)		# many 是否是多条数据，如果单条使用False
        for k in ser.data:
            for i in k["environment_set"]:
                k.setdefault("e_list",[])
                k['e_list'].append(i.split("|"))
        return Response(ser.data)		# ensure_ascii json时显示中文


class StandartPageNumberPagination(PageNumberPagination):
    page_size = 10 #默认每⻚返回的条数
    max_page_size = 20 #每⻚返回的最⼤条数
    page_size_query_param = 'page_size' #url中设置 page_size的键,默认为page_size
    page_query_param = 'page' #url中设置 page的键,默认为page

class Ip_info(APIView):
    def get(self,request,*args,**kwargs):
        env_name = request.GET.get("e")
        bus_name = request.GET.get("b")
        order = request.GET.get("order")
        filter_web = request.GET.get("filter")
        select_web = request.GET.get("select")
        if not order:
            order = 'id'
        if not filter_web:
            filter_web = ''

        print(env_name, bus_name, order,filter_web)
        times = request.GET.get('timestrap')
        if not times:
            times = datetime.datetime.now().strftime("%Y-%m-%d")
        bus = Environment.objects.get(name=env_name, to_b__name=bus_name)

        if not select_web:
            dic = {"env":bus,"times__pub_date":times,"name__contains":filter_web}
        else:
            dic = {"env": bus, "times__pub_date": times, "name__contains": filter_web,"type":select_web}

        if order == 'id':
            # ips = IpInfo.objects.filter(env=bus, times__pub_date=times,name__contains=filter_web).order_by(order)
            ips = IpInfo.objects.filter(**dic).order_by(order)
        else:
            # ips = IpInfo.objects.filter(env=bus, times__pub_date=times,name__contains=filter_web).extra(select={'num':order+"+0"})
            ips = IpInfo.objects.filter(**dic).extra(select={'num':order+"+0"})
            ips = ips.extra(order_by=["num"])
        # 分页
        pg = StandartPageNumberPagination()
        pg_ips = pg.paginate_queryset(queryset=ips, request=request, view=self)
        ipser = IpinfoSerializer(instance=pg_ips, many=True)

        totalcount = IpInfo.objects.filter(env=bus,times__pub_date=times).count()
        vircount = IpInfo.objects.filter(env=bus,times__pub_date=times,type=0).count()
        print(123)
        ip_dict = {
            "servertotalcount":totalcount,"vircount":vircount,"phycount":totalcount-vircount,
            'select_number':ips.count(),
            "IP":ipser.data
        }
        print(456)
        return Response(ip_dict)


class All_info(APIView):

    def get(self,request,*args,**kwargs):

        env_name = request.GET.get("e")
        bus_name = request.GET.get("b")
        envobj = Environment.objects.get(name=env_name, to_b__name=bus_name)
        print(env_name,bus_name,"all")
        timestrap = request.GET.get('timestrap')
        if not timestrap:
            timestrap = datetime.datetime.now().strftime("%Y-%m-%d")
        busobj = BusinessUnit.objects.get(id=envobj.to_b_id)
        # timobj = RecordingTime.objects.get(pub_date=timestrap)
        alldata = AllData.objects.filter(env=envobj,time__pub_date=timestrap)
        result = {
            "service":[0,0],
            "bigdata":[0,0]
        }
        service_n = 0
        big_n = 0
        ip_obj = IpInfo.objects.filter(times__pub_date=timestrap,env=envobj).values("cpurate","cpucore","type")
        for i in ip_obj:
            if i["type"] == 0:
                service_n += 1
                result["service"][0] += int(i["cpucore"])
                if not i["cpurate"] == '':
                    result["service"][1] += float(i["cpurate"])
                else:
                    result["service"][1] += float("0.00")
            else:
                big_n += 1
                result["bigdata"][0] += int(i["cpucore"])
                if not i["cpurate"] == '':
                    result["bigdata"][1] += float(i["cpurate"])
                else:
                    result["bigdata"][1] += float("0.00")
        if result["service"][1]>0:
            result["service"][1] = result["service"][1]/service_n

        if result["bigdata"][1] > 0:
            result["bigdata"][1] = result["bigdata"][1]/big_n
        print(result,big_n,service_n)
        alldataser = AlldataSerializer(instance=alldata,many=True)
        service = {
            'memorytotal': 0,
            'memoryuse': 0,
            'memoryrate': 0,
            'disktotal': 0,
            'diskuse': 0,
            'diskrate': 0,
            'cpurate': 0,
            "cpucore": 0
        }
        dic = {}
        for i in alldataser.data:
            if i['type'] == 0:
                i.setdefault("cpurate",float("%.2f"%result["service"][1]))
                i.setdefault("cpucore",float("%.2f"%result["service"][0]))
                dic['service'] = i
            else:
                i.setdefault("cpurate", float("%.2f"%result["bigdata"][1]))
                i.setdefault("cpucore", float("%.2f" % result["bigdata"][0]))
                dic['bigdata'] = i

        dic.setdefault('service', service)
        dic.setdefault('bigdata', service)
        print(dic)
        all_dict = {
            'time': timestrap,
            'business': busobj.name,
            'all': dic,
        }
        return Response(all_dict)

class Home(APIView):
    def get(self,request,*args,**kwargs):
        dic = {}
        ip_all = models.IpConfig.objects.all().count()
        e_all = models.Environment.objects.all().count()
        b_obj = models.BusinessUnit.objects.all()
        dic = {
            'ip_all':ip_all,
            'e_all':e_all,
            'number':{},
            'b_list':[
                ['product', '物理机', '虚拟机'],
            ],

        }
        for i in b_obj:
            v_number = models.IpConfig.objects.filter(env__to_b_id=i.id,servertype=0).count()
            s_number = models.IpConfig.objects.filter(env__to_b_id=i.id,servertype=1).count()
            name = i.name
            # dic['number'][name] = v_number + s_number
            dic['b_list'].append([name,s_number,v_number])

        print(dic)
        return Response(dic)




from django.http import FileResponse


class Download(APIView):

    def get(self, request,*args,**kwargs):

        timestrap = request.GET.get("timestrap")
        head_data = ['序号','IP地址','内存总量/M','内存使用量/M','内存使用率','cpu核数','平均load值','cpu使用率',
                     '磁盘总量/G','磁盘使用量/G','磁盘使用率','服务器类型','服务类型']

        alldict = {}
        # ips = None
        if timestrap:
            ips = IpInfo.objects.filter(times__pub_date='2019-12-02')
        else:
            ips = IpInfo.objects.filter(times__pub_date=datetime.datetime.now().strftime('%Y-%m-%d'))


        for j in ips:

            e = j.env.name
            b = j.env.to_b.name
            alldict.setdefault(b,{})
            alldict[b].setdefault(e,[])
            if j.get_type_display() == 'service':
                type_result = '服务'
            else:
                type_result = '大数据'
            alldict[b][e].append([len(alldict[b][e])+1,j.name,j.memorytotal,j.memoryuse,j.memoryrate+'%',
                                 j.cpucore,str(j.averageload),j.cpurate+'%',float(j.disktotal),float(j.diskuse),
                                  j.diskrate+'%',j.get_servertype_display(),type_result])
        # 样式设置
        alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(border_style="thin", color=BLACK)
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_title_font = Font(name='宋体', size=14, bold=True, color=BLACK)
        row_title_fill = PatternFill(fill_type='solid', fgColor=GREEN)
        content_font = Font(name='宋体', size=12, bold=False, color=BLACK)
        content_fill = PatternFill(fill_type='solid', fgColor=GREEN)
        title_font = Font(name="黑体", bold=True, size=16,color=BLACK)

        wb = Workbook()
        sheet = wb.active
        status = True

        for k1,v1 in alldict.items():
            if status:
                sheet.title = k1
                status = False
            else:
                sheet = wb.create_sheet(k1)
            num = 1
            for k2,v2 in v1.items():

                sheet.merge_cells(start_row=num, start_column=1, end_row=num, end_column=13)
                sheet.cell(num, 1).value = k2
                sheet.cell(num, 1).alignment = alignment
                sheet.cell(num, 1).font = title_font
                sheet.cell(num, 1).fill = row_title_fill
                num += 1
                sheet.append(head_data)
                rw = sheet.max_row
                sheet.column_dimensions['A'].width = 15
                sheet.column_dimensions['B'].width = 20
                sheet.column_dimensions['C'].width = 20
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width = 20
                sheet.column_dimensions['F'].width = 15
                sheet.column_dimensions['G'].width = 20
                sheet.column_dimensions['H'].width = 15
                sheet.column_dimensions['I'].width = 20
                sheet.column_dimensions['J'].width = 20
                sheet.column_dimensions['K'].width = 15
                sheet.column_dimensions['L'].width = 20
                sheet.column_dimensions['M'].width = 15

                for c in list(map(chr, range(ord('A'), ord('M') + 1))):

                    sheet[c+str(rw)].alignment = alignment
                    sheet[c+str(rw)].font = row_title_font


                num += 1
                vmnum = 0
                phnum = 0
                memtotal = 0
                memuse = 0
                disktotal = 0.0
                diskuse = 0.0
                for line in v2:
                    sheet.append(line)
                    memtotal += line[2]
                    memuse += line[3]
                    disktotal += line[8]
                    diskuse += line[9]
                    if '虚拟机' in line:
                        vmnum += 1
                    else:
                        phnum += 1
                    darw = sheet.max_row
                    for cc in list(map(chr, range(ord('A'), ord('M') + 1))):
                        pass
                        sheet[cc + str(darw)].alignment = alignment
                        sheet[cc + str(darw)].font = content_font
                    num += 1

                endrw = sheet.max_row

                # sheet.cell(row = endrw+1, column = 3).value = '=SUM(C'+str(rw+1)+':'+'C'+str(endrw)+')'
                sheet.cell(row = endrw+1, column = 3).value = memtotal
                sheet['C'+str(endrw+1)].font = content_font
                sheet['C'+str(endrw+1)].alignment = alignment

                # sheet.cell(row = endrw+1, column = 4).value = '=SUM(D'+str(rw+1)+':'+'D'+str(endrw)+')'
                sheet.cell(row = endrw+1, column = 4).value = memuse
                sheet['D' + str(endrw+1)].font = content_font
                sheet['D' + str(endrw+1)].alignment = alignment

                # sheet.cell(row = endrw+1, column = 8).value = '=SUM(H'+str(rw+1)+':'+'H'+str(endrw)+')'
                sheet.cell(row = endrw+1, column = 9).value = disktotal
                sheet['I' + str(endrw+1)].font = content_font
                sheet['I' + str(endrw+1)].alignment = alignment

                # sheet.cell(row = endrw+1, column = 9).value = '=SUM(I'+str(rw+1)+':'+'I'+str(endrw)+')'
                sheet.cell(row = endrw+1, column = 10).value = diskuse
                sheet['J' + str(endrw+1)].font = content_font
                sheet['J' + str(endrw+1)].alignment = alignment

                # 物理机 虚拟机    =COUNTIF(K2:K14,"虚拟机")
                sheet.cell(row=endrw + 1, column=12).value = str(vmnum) +'/'+ str(phnum)
                sheet['L' + str(endrw + 1)].font = content_font
                sheet['L' + str(endrw + 1)].alignment = alignment


                num += 1
                """
                   {'内存总量/M':{'C':3,}, '内存使用总量/M':{'D':4},'磁盘总量/G':{'I',9},'硬盘使用总量/G':{'J',10},'虚拟机/物理机':{'L',12} } 
                """

                sheet.cell(row=endrw + 2, column=3).value = '内存总量/M'
                sheet['C' + str(endrw+2)].font = content_font
                sheet['C' + str(endrw+2)].alignment = alignment
                sheet['C' + str(endrw+2)].fill = content_fill

                sheet.cell(row=endrw + 2, column=4).value = '内存使用总量/M'
                sheet['D' + str(endrw+2)].font = content_font
                sheet['D' + str(endrw+2)].alignment = alignment
                sheet['D' + str(endrw+2)].fill = content_fill


                sheet.cell(row=endrw + 2, column=9).value = '硬盘总量/G'
                sheet['I' + str(endrw+2)].font = content_font
                sheet['I' + str(endrw+2)].alignment = alignment
                sheet['I' + str(endrw+2)].fill = content_fill

                sheet.cell(row=endrw + 2, column=10).value = '硬盘使用总量/G'
                sheet['J' + str(endrw+2)].font = content_font
                sheet['J' + str(endrw+2)].alignment = alignment
                sheet['J' + str(endrw+2)].fill = content_fill

                sheet.cell(row=endrw + 2, column=12).value = '虚拟机/物理机'
                sheet['L' + str(endrw + 2)].font = content_font
                sheet['L' + str(endrw + 2)].alignment = alignment
                sheet['L' + str(endrw + 2)].fill = content_fill

                num += 1
                sheet.append([' ',' '])
                num += 1

        wb.save('ipdatas.xlsx')
        paths = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        os.path.join(paths, 'ipdatas.xlsx')
        file = open(os.path.join(paths, 'ipdatas.xlsx'), 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="ipdatas.xlsx"'

        return response

class Memory_table(APIView):
    def get(self,request,*args,**kwargs):
        env_name = request.GET.get("e")
        bus_name = request.GET.get("b")
        envobj = Environment.objects.get(name=env_name, to_b__name=bus_name)
        result = [[],[]]
        id_big = 0
        id_s = 0
        timestrap = request.GET.get('timestrap')
        if not timestrap:
            timestrap = datetime.datetime.now().strftime("%Y-%m-%d")
        ip_obj = models.IpInfo.objects.filter(times__pub_date=timestrap,env=envobj).extra(select={'num':"memoryrate"+"+0"}).extra(order_by=["num"]).values("memoryrate","type","name")
        for i in ip_obj:

            if i["type"] == 1:
                id_big += 1
                result[0].append([float(i["memoryrate"]),id_big,i["name"],"大数据"])
            else:
                id_s += 1
                result[1].append([float(i["memoryrate"]),id_s, i["name"], "服务"])

        return Response(result)


class Disk_table(APIView):
    def get(self,request,*args,**kwargs):
        env_name = request.GET.get("e")
        bus_name = request.GET.get("b")
        envobj = Environment.objects.get(name=env_name, to_b__name=bus_name)
        result = [[],[]]
        id_big = 0
        id_s = 0
        timestrap = request.GET.get('timestrap')
        if not timestrap:
            timestrap = datetime.datetime.now().strftime("%Y-%m-%d")
        ip_obj = models.IpInfo.objects.filter(times__pub_date=timestrap,env=envobj).extra(select={'num':"diskrate"+"+0"}).extra(order_by=["num"]).values("diskrate","type","name")
        for i in ip_obj:

            if i["type"] == 1:
                id_big += 1
                result[0].append([float(i["diskrate"]),id_big,i["name"],"大数据"])
            else:
                id_s += 1
                result[1].append([float(i["diskrate"]),id_s, i["name"], "服务"])

        return Response(result)


class Cpu_table(APIView):
    def get(self,request,*args,**kwargs):
        env_name = request.GET.get("e")
        bus_name = request.GET.get("b")
        envobj = Environment.objects.get(name=env_name, to_b__name=bus_name)
        result = [[],[]]
        id_big = 0
        id_s = 0
        timestrap = request.GET.get('timestrap')
        if not timestrap:
            timestrap = datetime.datetime.now().strftime("%Y-%m-%d")
        ip_obj = models.IpInfo.objects.filter(times__pub_date=timestrap,env=envobj).extra(select={'num':"cpurate"+"+0"}).extra(order_by=["num"]).values("cpurate","type","name")
        for i in ip_obj:
            if i["type"] == 1:
                id_big += 1
                result[0].append([0.00 if i["cpurate"] == '' else float(i["cpurate"]),id_big,i["name"],"大数据"])
            else:
                id_s += 1
                result[1].append([0.00 if i["cpurate"] == '' else float(i["cpurate"]), id_s, i["name"], "服务"])

        return Response(result)